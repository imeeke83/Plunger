# -*- coding: utf-8 -*-

import requests
from flask import Flask, request, Response
import math, sys
from konlpy.tag import Okt
import pickle
from re import split
from networkx import Graph
from networkx import pagerank
from itertools import combinations
from collections import Counter
from openpyxl import load_workbook

API_KEY = ''
app = Flask(__name__)

# User DataBase
EXCEL_FILE_NAME = r"C:\Users\KangmoonKo\sba\ComplaintDB.xlsx"
db = load_workbook(filename=EXCEL_FILE_NAME) # 기존에 존재하는 엑셀파일만 불러올 수 있음
userInfoDB = db['Complaint'] # db 엑셀파일 안의 userInfoDB 시트

#%%

class Sentence(object):
    okt = Okt()

    def __init__(self, text, index=0):
        self.index = index
        self.text = text.strip()
        self.tokens = self.okt.phrases(self.text)
        self.bow = Counter(self.tokens)

    def __str__(self):
        return self.text

    def __hash__(self):
        return self.index


class TextRank(object):
    def __init__(self, text):
        self.text = text.strip()
        self.build()

    def build(self):
        self._build_sentences()
        self._build_graph()
        self.pageranks = pagerank(self.graph, weight='weight')
        self.reordered = sorted(self.pageranks, key=self.pageranks.get, reverse=True)

    def _build_sentences(self):
        dup = {}
        candidates = split(r'(?:(?<=[^0-9])\.|\n)', self.text)
        self.sentences = []
        index = 0
        for candidate in candidates:
            while len(candidate) and (candidate[-1] == '.' or candidate[-1] == ' '):
                candidate = candidate.strip(' ').strip('.')
            if len(candidate) and candidate not in dup:
                dup[candidate] = True
                self.sentences.append(Sentence(candidate + '.', index))
                index += 1
        del dup
        del candidates

    def _build_graph(self):
        self.graph = Graph()
        self.graph.add_nodes_from(self.sentences)
        for sent1, sent2 in combinations(self.sentences, 2):
            weight = self._jaccard(sent1, sent2)
            if weight:
                self.graph.add_edge(sent1, sent2, weight=weight)

    def _jaccard(self, sent1, sent2):
        p = sum((sent1.bow & sent2.bow).values())
        q = sum((sent1.bow | sent2.bow).values())
        return p / q if q else 0

    def summarize(self, count=3, verbose=True):
        results = sorted(self.reordered[:count], key=lambda sentence: sentence.index)
        results = [result.text for result in results]
        if verbose:
            return '\n'.join(results)
        else:
            return results

#%%

class Filter:
    def __init__(self):
        self.words = set() #어떤 단어들이 있는지. 집합
        self.word_dict = {} #이중 dictionary; [카테고리][단어] 가 몇번 사용됬는지. 히스토그램
        self.category_dict = {} #[카테고리] 가 몇번 사용됬는지
    
    ## text를 조사 어미 구두점을 제외한 단어만 list로 반환
    def split(self, text):
        results = []
        twitter = Okt() #형태소 분석기
        malist = twitter.pos(text, norm=True, stem=True) #steam True로 text를 분석.
        
        # 실습 2
        # 아래 for 문을 한줄짜리 for 문으로 바꿔보세요 List Comprehension
        # for word in malist:
        #     if not word[1] in ["Josa", "Eomi","Punctuation"]:
        #         results.append(word[0])
        
        # 조사, 어미, 구두점 이 아닌 word에 대한 word만 result에 저장.
        results = [word[0] for word in malist if not word[1] in ["Josa","Eomi","Punctuation"]]
                
                
        return results

    ## word_dict 히스토램(word_dict)과, word 목록에 추가하는 작업.
    def inc_word(self, word, category):
        if not category in self.word_dict:
            self.word_dict[category] = {}
        if not word in self.word_dict[category]:
            self.word_dict[category][word] = 0
        self.word_dict[category][word] += 1
        self.words.add(word)

    ## 카테고리 히스토그램 만들기.
    def inc_category(self, category):
        if not category in self.category_dict:
            self.category_dict[category] = 0
        self.category_dict[category] += 1

    ## 텍스트 넣어서 histogram 만들기
    def fit(self, text, category):
        word_list = self.split(text) ## 조사 어미 구두점 제외하여 list로 반환
        for word in word_list:
            self.inc_word(word,category)
        self.inc_category(category)
    
    ## score를 확률적 계산
    ## P(카테고리|전체문서) + P( 단어 | 해당카테고리)
    def score(self, words, category):
        score = math.log(self.category_prob(category)) #해당 카테고리가 나올 확률
        for word in words: # 각 단어에 대한 확률의 합.
            score += math.log(self.word_prob(word, category))
        return score

    def predict(self, text):
        best_category = None
        max_score = -sys.maxsize
        words = self.split(text) #형태소 분석 (조사 어미 구두점 빼고 단어 list)
        score_list = [] # [(카테고리,score) ...] 쌍으로 들어감. socre_list
        for category in self.category_dict.keys():
            score = self.score(words, category)
            score_list.append((category,score))
                        
            if score > max_score: #가장 높은 score와 카테고리를 저장.
                max_score = score
                best_category = category
                
        return best_category, score_list
    
    ## 해당 단어가, 카테고리에서 몇번이나 쓰였는지 가져오는 함수.
    def get_word_count(self, word, category):
        if word in self.word_dict[category]:
            return self.word_dict[category][word]
        else:
            return 0
    
    ## 전체 문서수에 대해 해당 카테고리가 몇번이나 나왔는지. 확률. #카테고리가 나올 확률
    def category_prob(self, category):
        sum_categories = sum(self.category_dict.values()) # 전체 문서의 숫자
        category_v = self.category_dict[category] # 해당 카테고리의 숫자
        return category_v / sum_categories # 카테고리 수 / 전체 문서의 수
    
    ## 
    def word_prob(self, word, category): # 
        n = self.get_word_count(word, category) + 1 # 해당 단어가 카테고리에서 몇번이나 쓰였는지. log(0)이 없으므로 +1 로 bias
        d = sum(self.word_dict[category].values()) + len(self.words) # 해당 카테고리의 단어의 수 + 전체 단어의 수
        return n/d

    def save_as_pickle(self,file_name='TF-IDF'):
      with open(file_name+'.pickle', 'wb') as f:
          pickle.dump(self, f)

#%%
def parse_message(message):
    """
    telegram 에서 data 인자를 받아옴
    data 내부 구조를 이해해야 한다.
    
    Retuen :    
    chat_id = 사용자 아이디 코드
    msg = 사용자 대화 내용    
    """
    chat_id = message['message']['chat']['id']
    msg = message['message']['text']
    
    return chat_id, msg

# 사용자에게 메세지를 보냄
def send_message(user_id, text):
    """
    사용자에게 메세지를 보냄
    """
    print('send_message')
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
    params = {'chat_id': user_id, 'text': text}

    response = requests.post(url, json=params)
    return response
    
# 경로 설정, URL 설정
@app.route('/', methods=['POST', 'GET'])
def index():
    if request.method == 'POST':
        message = request.get_json()
        chat_id, msg = parse_message(message)
        
        if '안녕' in msg:
            send_message(chat_id, '안녕하세요. 저는 뚫어뻥입니다. 불편사항을 사이다처럼 시원하고 신속히 도와드릴게요 !')
            return Response('ok', status=200)
        elif '이름' in msg:
            send_message(chat_id, '신고 장소를 말씀해주세요.')
            userInfoDB[userInfoDB.max_row][0].value = msg
            db.save(EXCEL_FILE_NAME)
            return Response('ok', status=200)
        elif '장소' in msg:
            send_message(chat_id, '휴대폰 번호를 입력해주세요')
            userInfoDB[userInfoDB.max_row][2].value = msg
            db.save(EXCEL_FILE_NAME)
            return Response('ok', status=200)
        elif '연락처' in msg:
            send_message(chat_id, '민원신청을 위한 개인정보 수집 및 이용, 개인정보 취급 위탁에 동의하시나요?')
            userInfoDB[userInfoDB.max_row][1].value = msg
            db.save(EXCEL_FILE_NAME)
            return Response('ok', status=200)
        elif '응' in msg:
            send_message(chat_id, '접수가 완료되었습니다. 접수번호가 나오면 다시 알려드릴께요 !')
            return Response('ok', status=200)
        elif '접수' in msg[:2]:
            df = pickle.load(open(r'C:\Users\KangmoonKo\sba\intent_model.pickle','rb'))
            intent, scorelist = df.predict(msg)
            userInfoDB[userInfoDB.max_row+1][3].value = intent
            print("결과 =", intent)
            print(scorelist)
            send_message(chat_id, intent+'관련해서 불편하셨군요.\n민원접수를 바로 진행해드리겠습니다 !')
            textrank = TextRank(msg)
            textList = textrank.summarize(3, verbose=False)
            resultText = ''
            for i in textList:
                resultText = resultText + i + '\n'
            userInfoDB[userInfoDB.max_row][4].value = resultText
            send_message(chat_id, resultText+'\n 위와 같이 민원 내용을 정리해 전달해드리겠습니다 !')
            send_message(chat_id, '빠른 접수를 위해 성함을 알려주세요.')
            db.save(EXCEL_FILE_NAME)
            return Response('ok', status=200)
            
        return Response('ok', status=200)
    else:
        return 'Hello World!'


@app.route('/about')
def about():
  return 'About page'


if __name__ == '__main__':
    app.run(port = 5000)

